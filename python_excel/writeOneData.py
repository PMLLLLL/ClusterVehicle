import xlwings as xw  # 导入xlwings模块
import pandas as pd
import re
from openpyxl import load_workbook
from pathlib import Path

# 指定文件路径
file_path = Path('writeOneData.xlsx')

# 检查文件是否存在
if file_path.is_file():
    print(f"文件 '{file_path}' 存在！")
else:
    print(f"文件 '{file_path}' 不存在！")
    app = xw.App(visible = True, add_book = False)  # 启动Excel程序，不新建工作簿
    workbook = app.books.add()  # 新建工作簿
    print(f"创建文件 '{file_path}'！")
    workbook.save(f'{file_path}')  # 保存新建的多个工作簿

    app.quit()

# 读取 Excel 文件到 DataFrame
df = pd.read_excel(file_path,sheet_name=0, engine='openpyxl',header=None)

# 修改 DataFrame 中的特定单元格（假设我们要修改第 3 行第 2 列，即 B3）
df.at[2, 0] = -1233



# 加载 Excel 文件
# wb = load_workbook(file_path)

# # 如果 `Sheet1` 存在，删除它并重新创建
# if 'Sheet1' in wb.sheetnames:
#     del wb['Sheet1']
#     wb.create_sheet('Sheet1')  # 重新创建空白 Sheet1
#
# wb.save(file_path)
# wb.close()

# 将修改后的 DataFrame 写回 Excel 文件
df.to_excel(file_path, index=False, header=False)
print(f"写入数据成功！")
