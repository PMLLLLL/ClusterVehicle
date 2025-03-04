import pandas as pd
import re
from openpyxl import load_workbook

# 在数据表中找到指定车牌号数据
# 1.判断数据是否足够
# 2.将收集到的数据合并
# 3.输出到一个excel中

# 读取 Excel 文件
df = pd.read_excel("原始数据流.xlsx", sheet_name=0, engine='openpyxl')

# 输出 DataFrame 形状
print(f"DataFrame 共有 {df.shape[0]} 行 和 {df.shape[1]} 列")

# 记录保存的数据有几个
dataSaveNum = 0
series_data = []
data = pd.Series()  # 转换为 Series
newData = pd.Series()  # 转换为 Series
empty_series = pd.Series() #空行

# 获取特定单元格数据，例如获取第一行第一列的数据
for i in range(df.shape[0]):  # 0 到 82950
    cell_value = df.iloc[i, 0] # 行索引i，列索引0
    # 确保 cell_value 是字符串类型，处理 NaN 和非字符串类型
    cell_value = str(cell_value)  # 将单元格值转换为字符串
    # 使用 re（正则表达式，适用于复杂匹配）来检测str内容
    pattern = r".*C.*2380.*"  # 确保 C 在 2380 之前
    if re.match(pattern, cell_value):
        print(f"第 {i+1} 行字符串符合要求: {cell_value}")
        # 计算检测到的数据个数
        testCount = 0
        for j in range(70):  # 0 到 82950
            if pd.isna(df.iloc[i, j]) or df.iloc[i, j] == "":
                testCount = testCount
            else:
                testCount += 1
        if testCount >= 7:
            data = df.iloc[i:i+17]
            if dataSaveNum == 0:
                newData = data
            else:
                newData = pd.concat([newData, empty_series.to_frame().T], axis=0)
                newData = pd.concat([newData,data], axis=0)
            dataSaveNum += 1  # 记录保存的数据+1个
            print(f"第 {dataSaveNum} 次数据找到,数据有效个数为 {testCount}")


# 设定要写入的 Excel 文件
excel_path = "output.xlsx"

# 加载 Excel 文件
wb = load_workbook(excel_path)

# 如果 `Sheet1` 存在，删除它并重新创建
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']
    wb.create_sheet('Sheet1')  # 重新创建空白 Sheet1

wb.save(excel_path)
wb.close()

# 使用 ExcelWriter 写入
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    newData.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False, header=False)

print(f"数据已成功写入 {excel_path} 的第 1 行")




